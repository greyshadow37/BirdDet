import os
import json
import time
import glob
from pathlib import Path
from collections import defaultdict
import numpy as np
import pandas as pd


def calculate_iou_xyxy(box1, box2):
    """
    Calculate IoU between two boxes in [x1, y1, x2, y2] format.
    """
    x1 = max(box1[0], box2[0])
    y1 = max(box1[1], box2[1])
    x2 = min(box1[2], box2[2])
    y2 = min(box1[3], box2[3])

    inter_w = max(0.0, x2 - x1)
    inter_h = max(0.0, y2 - y1)
    inter_area = inter_w * inter_h

    area1 = max(0.0, box1[2] - box1[0]) * max(0.0, box1[3] - box1[1])
    area2 = max(0.0, box2[2] - box2[0]) * max(0.0, box2[3] - box2[1])
    union_area = area1 + area2 - inter_area

    if union_area <= 0:
        return 0.0
    return float(inter_area / union_area)


def compute_ap_from_matches(tp_list, conf_list, num_gts):
    """
    Compute 101-point interpolated Average Precision (AP) for a sequence of predictions.
    tp_list: list of 1s (TP) and 0s (FP) sorted in descending order of confidence.
    conf_list: list of confidence scores.
    num_gts: total number of ground-truth objects.
    """
    if num_gts == 0:
        return 1.0 if len(tp_list) == 0 else 0.0
    if len(tp_list) == 0:
        return 0.0

    tps = np.array(tp_list)
    fps = 1 - tps

    tp_cumsum = np.cumsum(tps)
    fp_cumsum = np.cumsum(fps)

    recalls = tp_cumsum / float(num_gts)
    precisions = tp_cumsum / (tp_cumsum + fp_cumsum + 1e-16)

    # 101-point interpolated AP (COCO standard)
    recall_thresholds = np.linspace(0.0, 1.0, 101)
    precisions_interp = []

    for r in recall_thresholds:
        prec_above_r = precisions[recalls >= r]
        if len(prec_above_r) > 0:
            precisions_interp.append(np.max(prec_above_r))
        else:
            precisions_interp.append(0.0)

    return float(np.mean(precisions_interp))


def match_detections(gts, preds, iou_thresh=0.5):
    """
    Match predictions to ground truth objects for a single image.
    gts: list of dicts with keys: 'class_id', 'bbox' ([x1, y1, x2, y2])
    preds: list of dicts with keys: 'class_id', 'bbox' ([x1, y1, x2, y2]), 'score'

    Returns:
        tp_count, fp_count, fn_count, matched_ious, per_pred_matches
        where per_pred_matches is a list of (score, class_id, is_tp)
    """
    # Sort predictions by descending confidence
    sorted_preds = sorted(preds, key=lambda x: x.get('score', 0.0), reverse=True)

    matched_gt_indices = set()
    per_pred_matches = []
    matched_ious = []

    for p in sorted_preds:
        p_cls = p['class_id']
        p_box = p['bbox']
        p_score = p.get('score', 1.0)

        best_iou = 0.0
        best_gt_idx = -1

        for idx, gt in enumerate(gts):
            if gt['class_id'] == p_cls and idx not in matched_gt_indices:
                iou = calculate_iou_xyxy(p_box, gt['bbox'])
                if iou > best_iou:
                    best_iou = iou
                    best_gt_idx = idx

        if best_iou >= iou_thresh and best_gt_idx != -1:
            matched_gt_indices.add(best_gt_idx)
            matched_ious.append(best_iou)
            per_pred_matches.append((p_score, p_cls, 1, best_iou))
        else:
            per_pred_matches.append((p_score, p_cls, 0, best_iou))

    tp_count = len(matched_gt_indices)
    fp_count = len(sorted_preds) - tp_count
    fn_count = len(gts) - tp_count

    return tp_count, fp_count, fn_count, matched_ious, per_pred_matches


def evaluate_single_image(image_path, width, height, gts, preds, class_names=None, conf_thresh=0.25, iou_thresh=0.5, latency_ms=0.0):
    """
    Evaluate detections for a single image and return a detailed metrics dictionary.
    """
    image_name = os.path.basename(image_path)
    
    # Filter predictions by confidence threshold
    filtered_preds = [p for p in preds if p.get('score', 1.0) >= conf_thresh]
    
    tp_count, fp_count, fn_count, matched_ious, pred_matches = match_detections(
        gts, filtered_preds, iou_thresh=iou_thresh
    )

    gt_count = len(gts)
    pred_count = len(filtered_preds)

    # Precision, Recall, F1
    if pred_count > 0:
        precision = tp_count / float(pred_count)
    else:
        precision = 1.0 if gt_count == 0 else 0.0

    if gt_count > 0:
        recall = tp_count / float(gt_count)
    else:
        recall = 1.0 if pred_count == 0 else 0.0

    if (precision + recall) > 0:
        f1_score = 2.0 * (precision * recall) / (precision + recall)
    else:
        f1_score = 0.0

    mean_iou = float(np.mean(matched_ious)) if len(matched_ious) > 0 else (1.0 if (gt_count == 0 and pred_count == 0) else 0.0)

    # Calculate AP@50, AP@75, and AP@50:95 for this image
    sorted_filtered_preds = sorted(filtered_preds, key=lambda x: x.get('score', 0.0), reverse=True)
    
    # AP@50
    tp_50_list = [match[2] for match in pred_matches]
    conf_list = [match[0] for match in pred_matches]
    ap_50 = compute_ap_from_matches(tp_50_list, conf_list, gt_count)

    # AP@75
    _, _, _, _, pred_matches_75 = match_detections(gts, filtered_preds, iou_thresh=0.75)
    tp_75_list = [m[2] for m in pred_matches_75]
    ap_75 = compute_ap_from_matches(tp_75_list, conf_list, gt_count)

    # AP@50:95 (average over IoU thresholds 0.50 to 0.95 with step 0.05)
    iou_thresholds = np.linspace(0.50, 0.95, 10)
    ap_list = []
    for t in iou_thresholds:
        _, _, _, _, t_matches = match_detections(gts, filtered_preds, iou_thresh=t)
        t_tp = [m[2] for m in t_matches]
        ap_list.append(compute_ap_from_matches(t_tp, conf_list, gt_count))
    ap_50_95 = float(np.mean(ap_list))

    # Ground truth and predicted class breakdown
    def get_class_name(cid):
        if class_names:
            if isinstance(class_names, dict):
                return str(class_names.get(cid, class_names.get(str(cid), cid)))
            elif isinstance(class_names, list) and 0 <= cid < len(class_names):
                return str(class_names[cid])
        return f"Class_{cid}"

    gt_cls_counts = defaultdict(int)
    for gt in gts:
        gt_cls_counts[get_class_name(gt['class_id'])] += 1
    gt_classes_str = ", ".join([f"{k}: {v}" for k, v in gt_cls_counts.items()]) if gt_cls_counts else "None"

    pred_cls_counts = defaultdict(int)
    pred_conf_list = []
    for p in filtered_preds:
        pred_cls_counts[get_class_name(p['class_id'])] += 1
        pred_conf_list.append(p.get('score', 1.0))
    pred_classes_str = ", ".join([f"{k}: {v}" for k, v in pred_cls_counts.items()]) if pred_cls_counts else "None"

    mean_conf = float(np.mean(pred_conf_list)) if pred_conf_list else 0.0

    return {
        "Image Name": image_name,
        "Image Path": str(image_path),
        "Width": int(width) if width else 0,
        "Height": int(height) if height else 0,
        "GT Count": gt_count,
        "Pred Count": pred_count,
        "TP": tp_count,
        "FP": fp_count,
        "FN": fn_count,
        "Precision": round(precision, 4),
        "Recall": round(recall, 4),
        "F1 Score": round(f1_score, 4),
        "Mean IoU": round(mean_iou, 4),
        "AP@50": round(ap_50, 4),
        "AP@75": round(ap_75, 4),
        "AP@50:95": round(ap_50_95, 4),
        "Mean Confidence": round(mean_conf, 4),
        "GT Classes": gt_classes_str,
        "Predicted Classes": pred_classes_str,
        "Inference Latency (ms)": round(latency_ms, 2)
    }


def compute_dataset_summary_and_classes(per_image_records, all_gts_by_image, all_preds_by_image, class_names=None, conf_thresh=0.25, iou_thresh=0.5):
    """
    Compute dataset-level overall summary and per-class metrics breakdown.
    """
    total_images = len(per_image_records)
    total_gt = sum(r["GT Count"] for r in per_image_records)
    total_preds = sum(r["Pred Count"] for r in per_image_records)
    total_tp = sum(r["TP"] for r in per_image_records)
    total_fp = sum(r["FP"] for r in per_image_records)
    total_fn = sum(r["FN"] for r in per_image_records)

    global_precision = total_tp / float(total_preds) if total_preds > 0 else (1.0 if total_gt == 0 else 0.0)
    global_recall = total_tp / float(total_gt) if total_gt > 0 else (1.0 if total_preds == 0 else 0.0)
    global_f1 = (2 * global_precision * global_recall / (global_precision + global_recall)) if (global_precision + global_recall) > 0 else 0.0

    mean_img_precision = float(np.mean([r["Precision"] for r in per_image_records])) if per_image_records else 0.0
    mean_img_recall = float(np.mean([r["Recall"] for r in per_image_records])) if per_image_records else 0.0
    mean_img_f1 = float(np.mean([r["F1 Score"] for r in per_image_records])) if per_image_records else 0.0
    mean_iou = float(np.mean([r["Mean IoU"] for r in per_image_records])) if per_image_records else 0.0
    
    mean_ap50 = float(np.mean([r["AP@50"] for r in per_image_records])) if per_image_records else 0.0
    mean_ap75 = float(np.mean([r["AP@75"] for r in per_image_records])) if per_image_records else 0.0
    mean_map50_95 = float(np.mean([r["AP@50:95"] for r in per_image_records])) if per_image_records else 0.0

    latencies = [r["Inference Latency (ms)"] for r in per_image_records if r["Inference Latency (ms)"] > 0]
    avg_latency = float(np.mean(latencies)) if latencies else 0.0
    fps = round(1000.0 / avg_latency, 2) if avg_latency > 0 else 0.0

    summary_data = [
        {"Metric": "Total Images Evaluated", "Value": total_images},
        {"Metric": "Total Ground Truth Objects", "Value": total_gt},
        {"Metric": "Total Predictions (>= conf threshold)", "Value": total_preds},
        {"Metric": "Total True Positives (TP)", "Value": total_tp},
        {"Metric": "Total False Positives (FP)", "Value": total_fp},
        {"Metric": "Total False Negatives (FN)", "Value": total_fn},
        {"Metric": "Global Precision", "Value": round(global_precision, 4)},
        {"Metric": "Global Recall", "Value": round(global_recall, 4)},
        {"Metric": "Global F1 Score", "Value": round(global_f1, 4)},
        {"Metric": "Mean Image Precision", "Value": round(mean_img_precision, 4)},
        {"Metric": "Mean Image Recall", "Value": round(mean_img_recall, 4)},
        {"Metric": "Mean Image F1 Score", "Value": round(mean_img_f1, 4)},
        {"Metric": "Mean IoU", "Value": round(mean_iou, 4)},
        {"Metric": "mAP@50", "Value": round(mean_ap50, 4)},
        {"Metric": "mAP@75", "Value": round(mean_ap75, 4)},
        {"Metric": "mAP@50:95", "Value": round(mean_map50_95, 4)},
        {"Metric": "Average Latency per Image (ms)", "Value": round(avg_latency, 2)},
        {"Metric": "Throughput (FPS)", "Value": fps},
        {"Metric": "Confidence Threshold", "Value": conf_thresh},
        {"Metric": "IoU Threshold", "Value": iou_thresh}
    ]

    # Compute Per-Class Breakdown
    # Collect all unique class IDs
    unique_class_ids = set()
    for gts in all_gts_by_image.values():
        for g in gts:
            unique_class_ids.add(g['class_id'])
    for preds in all_preds_by_image.values():
        for p in preds:
            if p.get('score', 1.0) >= conf_thresh:
                unique_class_ids.add(p['class_id'])

    if class_names:
        if isinstance(class_names, dict):
            for k in class_names.keys():
                try:
                    unique_class_ids.add(int(k))
                except ValueError:
                    pass
        elif isinstance(class_names, list):
            for idx in range(len(class_names)):
                unique_class_ids.add(idx)

    def get_class_name(cid):
        if class_names:
            if isinstance(class_names, dict):
                return str(class_names.get(cid, class_names.get(str(cid), cid)))
            elif isinstance(class_names, list) and 0 <= cid < len(class_names):
                return str(class_names[cid])
        return f"Class_{cid}"

    per_class_records = []
    iou_thresholds = np.linspace(0.50, 0.95, 10)

    for cid in sorted(unique_class_ids):
        cname = get_class_name(cid)
        cls_gt_count = 0
        cls_pred_count = 0
        cls_tp = 0
        cls_fp = 0
        cls_matched_ious = []
        
        cls_pred_eval_list = [] # (score, is_tp)
        cls_pred_eval_75_list = []
        cls_pred_eval_multi = {t: [] for t in iou_thresholds}

        for img_name in all_gts_by_image.keys():
            img_gts = [g for g in all_gts_by_image.get(img_name, []) if g['class_id'] == cid]
            img_preds = [p for p in all_preds_by_image.get(img_name, []) if p['class_id'] == cid and p.get('score', 1.0) >= conf_thresh]
            
            cls_gt_count += len(img_gts)
            cls_pred_count += len(img_preds)

            # Match at 0.50
            tp, fp, fn, matched_ious, matches = match_detections(img_gts, img_preds, iou_thresh=0.50)
            cls_tp += tp
            cls_fp += fp
            cls_matched_ious.extend(matched_ious)
            for m in matches:
                cls_pred_eval_list.append((m[0], m[2]))

            # Match at 0.75
            _, _, _, _, matches_75 = match_detections(img_gts, img_preds, iou_thresh=0.75)
            for m in matches_75:
                cls_pred_eval_75_list.append((m[0], m[2]))

            # Match multi
            for t in iou_thresholds:
                _, _, _, _, m_t = match_detections(img_gts, img_preds, iou_thresh=t)
                for m in m_t:
                    cls_pred_eval_multi[t].append((m[0], m[2]))

        cls_fn = cls_gt_count - cls_tp
        cls_precision = cls_tp / float(cls_pred_count) if cls_pred_count > 0 else (1.0 if cls_gt_count == 0 else 0.0)
        cls_recall = cls_tp / float(cls_gt_count) if cls_gt_count > 0 else (1.0 if cls_pred_count == 0 else 0.0)
        cls_f1 = (2 * cls_precision * cls_recall / (cls_precision + cls_recall)) if (cls_precision + cls_recall) > 0 else 0.0
        cls_mean_iou = float(np.mean(cls_matched_ious)) if cls_matched_ious else 0.0

        # Sort all class preds across images by score for AP computation
        cls_pred_eval_list.sort(key=lambda x: x[0], reverse=True)
        cls_tp_50 = [x[1] for x in cls_pred_eval_list]
        cls_conf_50 = [x[0] for x in cls_pred_eval_list]
        cls_ap50 = compute_ap_from_matches(cls_tp_50, cls_conf_50, cls_gt_count)

        cls_pred_eval_75_list.sort(key=lambda x: x[0], reverse=True)
        cls_tp_75 = [x[1] for x in cls_pred_eval_75_list]
        cls_conf_75 = [x[0] for x in cls_pred_eval_75_list]
        cls_ap75 = compute_ap_from_matches(cls_tp_75, cls_conf_75, cls_gt_count)

        cls_ap_multi = []
        for t in iou_thresholds:
            cls_pred_eval_multi[t].sort(key=lambda x: x[0], reverse=True)
            t_tp = [x[1] for x in cls_pred_eval_multi[t]]
            t_conf = [x[0] for x in cls_pred_eval_multi[t]]
            cls_ap_multi.append(compute_ap_from_matches(t_tp, t_conf, cls_gt_count))
        cls_ap50_95 = float(np.mean(cls_ap_multi))

        per_class_records.append({
            "Class ID": cid,
            "Class Name": cname,
            "GT Count": cls_gt_count,
            "Pred Count": cls_pred_count,
            "TP": cls_tp,
            "FP": cls_fp,
            "FN": cls_fn,
            "Precision": round(cls_precision, 4),
            "Recall": round(cls_recall, 4),
            "F1 Score": round(cls_f1, 4),
            "Mean IoU": round(cls_mean_iou, 4),
            "AP@50": round(cls_ap50, 4),
            "AP@75": round(cls_ap75, 4),
            "AP@50:95": round(cls_ap50_95, 4)
        })

    return summary_data, per_class_records


def export_metrics_to_excel(per_image_records, summary_data, per_class_records, excel_filepath):
    """
    Save metrics into a multi-sheet Excel file (.xlsx) and a companion CSV file.
    """
    out_dir = os.path.dirname(os.path.abspath(excel_filepath))
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    df_per_image = pd.DataFrame(per_image_records)
    df_summary = pd.DataFrame(summary_data)
    df_per_class = pd.DataFrame(per_class_records)

    # Export Excel (.xlsx) with styled formatting
    with pd.ExcelWriter(excel_filepath, engine='openpyxl') as writer:
        df_per_image.to_excel(writer, sheet_name='Per-Image Metrics', index=False)
        df_summary.to_excel(writer, sheet_name='Summary Metrics', index=False)
        df_per_class.to_excel(writer, sheet_name='Per-Class Metrics', index=False)

        # Style worksheets using openpyxl
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )

            for ws in writer.book.worksheets:
                ws.views.sheetView[0].showGridLines = True
                for col_idx, col in enumerate(ws.columns, 1):
                    header_cell = col[0]
                    header_cell.fill = header_fill
                    header_cell.font = header_font
                    header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

                    for cell in col[1:]:
                        cell.border = thin_border
                        if isinstance(cell.value, (int, np.integer)):
                            cell.alignment = Alignment(horizontal="right")
                        elif isinstance(cell.value, (float, np.floating)):
                            cell.alignment = Alignment(horizontal="right")
                            cell.number_format = '0.0000'
        except Exception as e:
            # Fallback if styling fails
            pass

    # Also save companion CSV for per-image metrics
    csv_filepath = os.path.splitext(excel_filepath)[0] + ".csv"
    df_per_image.to_csv(csv_filepath, index=False)

    return excel_filepath, csv_filepath


def load_yolo_ground_truths(images_dir, labels_dir=None, image_files=None):
    """
    Load YOLO format ground-truth bounding boxes for a directory of images.
    Returns: dict mapping image_filename -> list of {'class_id': int, 'bbox': [x1, y1, x2, y2]}
    """
    import cv2

    if image_files is None:
        image_extensions = ('*.jpg', '*.jpeg', '*.png', '*.JPG', '*.JPEG', '*.PNG', '*.bmp')
        image_files = []
        for ext in image_extensions:
            image_files.extend(glob.glob(os.path.join(images_dir, ext)))

    gts_by_image = {}
    image_shapes = {}

    for img_path in sorted(image_files):
        img_name = os.path.basename(img_path)
        stem = Path(img_path).stem

        # Try to find corresponding label file
        label_file = None
        candidate_paths = []
        if labels_dir and os.path.isdir(labels_dir):
            candidate_paths.append(os.path.join(labels_dir, f"{stem}.txt"))
        
        # Standard YOLO pattern: replace 'images' with 'labels' in image path
        norm_img_path = os.path.normpath(img_path)
        if "images" in norm_img_path:
            candidate_paths.append(norm_img_path.replace("images", "labels").rsplit(".", 1)[0] + ".txt")

        # Check adjacent labels directory
        parent_dir = os.path.dirname(img_path)
        candidate_paths.append(os.path.join(parent_dir, "..", "labels", f"{stem}.txt"))
        candidate_paths.append(os.path.join(parent_dir, "..", "labels", "test", f"{stem}.txt"))
        candidate_paths.append(os.path.join(parent_dir, f"{stem}.txt"))

        for cand in candidate_paths:
            if os.path.isfile(cand):
                label_file = cand
                break

        # Read image dimensions
        img = cv2.imread(img_path)
        if img is not None:
            h, w = img.shape[:2]
        else:
            w, h = 512, 512
        image_shapes[img_name] = (w, h)

        gts = []
        if label_file and os.path.isfile(label_file):
            try:
                with open(label_file, 'r', encoding='utf-8') as f:
                    for line in f:
                        parts = line.strip().split()
                        if len(parts) >= 5:
                            cls_id = int(float(parts[0]))
                            xc, yc, bw, bh = map(float, parts[1:5])
                            # Convert normalized xywh to absolute xyxy
                            x1 = (xc - bw / 2.0) * w
                            y1 = (yc - bh / 2.0) * h
                            x2 = (xc + bw / 2.0) * w
                            y2 = (yc + bh / 2.0) * h
                            gts.append({
                                'class_id': cls_id,
                                'bbox': [max(0.0, x1), max(0.0, y1), min(float(w), x2), min(float(h), y2)]
                            })
            except Exception as e:
                print(f"Warning: Failed reading label file {label_file}: {e}")

        gts_by_image[img_name] = gts

    return gts_by_image, image_shapes


def load_coco_ground_truths(ann_json_path, images_dir=None):
    """
    Load COCO format annotations JSON.
    Returns:
        gts_by_image: dict mapping image_filename -> list of {'class_id': int, 'bbox': [x1, y1, x2, y2]}
        image_shapes: dict mapping image_filename -> (width, height)
        categories: dict mapping class_id -> class_name
    """
    if not os.path.isfile(ann_json_path):
        raise FileNotFoundError(f"COCO annotation file not found: {ann_json_path}")

    with open(ann_json_path, 'r', encoding='utf-8') as f:
        coco = json.load(f)

    categories = {cat['id']: cat['name'] for cat in coco.get('categories', [])}

    # Map image id to filename and shape
    img_id_to_info = {}
    image_shapes = {}
    gts_by_image = defaultdict(list)

    for img in coco.get('images', []):
        img_id = img['id']
        fn = os.path.basename(img['file_name'])
        w = img.get('width', 0)
        h = img.get('height', 0)
        img_id_to_info[img_id] = fn
        image_shapes[fn] = (w, h)
        if fn not in gts_by_image:
            gts_by_image[fn] = []

    for ann in coco.get('annotations', []):
        img_id = ann['image_id']
        if img_id in img_id_to_info:
            fn = img_id_to_info[img_id]
            cid = ann['category_id']
            # COCO bbox: [x, y, width, height]
            bx, by, bw, bh = ann['bbox']
            x1 = bx
            y1 = by
            x2 = bx + bw
            y2 = by + bh
            gts_by_image[fn].append({
                'class_id': cid,
                'bbox': [float(x1), float(y1), float(x2), float(y2)]
            })

    return dict(gts_by_image), image_shapes, categories
